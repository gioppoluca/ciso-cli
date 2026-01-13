from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook


@dataclass(frozen=True)
class TableReader:
    """
    Read tabular data from CSV or Excel into a list[dict].

    - CSV: uses utf-8-sig to handle BOM
    - XLSX: reads the first sheet by default (or a named sheet)
    - Headers are normalized by stripping spaces
    """

    def read(self, file: Path, sheet: Optional[str] = None) -> list[dict[str, Any]]:
        ext = file.suffix.lower()
        if ext == ".csv":
            return self._read_csv(file)
        if ext in (".xlsx", ".xlsm"):
            return self._read_xlsx(file, sheet=sheet)
        raise ValueError(f"Unsupported file type: {ext}. Use .csv or .xlsx/.xlsm")

    def _read_csv(self, path: Path) -> list[dict[str, Any]]:
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            # normalize headers
            if reader.fieldnames:
                reader.fieldnames = [self._norm_header(h) for h in reader.fieldnames]
            out: list[dict[str, Any]] = []
            for row in reader:
                out.append({self._norm_header(k): v for k, v in row.items()})
            return out

    def _read_xlsx(self, path: Path, sheet: Optional[str]) -> list[dict[str, Any]]:
        wb = load_workbook(filename=path, read_only=True, data_only=True)
        ws = wb[sheet] if sheet else wb.worksheets[0]
        rows = ws.iter_rows(values_only=True)

        headers = next(rows, None)
        if not headers:
            return []

        headers_norm = [self._norm_header(h) for h in headers]

        out: list[dict[str, Any]] = []
        for r in rows:
            d: dict[str, Any] = {}
            for i in range(min(len(headers_norm), len(r))):
                key = headers_norm[i]
                if key:  # ignore empty header cells
                    d[key] = r[i]
            out.append(d)
        return out

    def _norm_header(self, h: Any) -> str:
        if h is None:
            return ""
        return str(h).strip()
